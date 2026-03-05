"""
Genetics Cloud — Export Engine
===============================
PDF and Excel generation for schedules.
"""

import io
import logging
from datetime import datetime, time
from typing import List, Dict, Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scheduling.models import TimetableVersion

logger = logging.getLogger(__name__)


# ═════════════════════════════════════════════════════════════
# PDF EXPORT
# ═════════════════════════════════════════════════════════════

class PDFExporter:
    """
    Generate PDF timetable reports.
    Supports: student schedule, lecturer dashboard, admin master pulse.
    """

    def __init__(self, title: str = "UniSched Report"):
        self.title = title
        self.styles = getSampleStyleSheet()
        self._add_custom_styles()

    def _add_custom_styles(self):
        """Add custom paragraph styles."""
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=self.styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#1E40AF"),
            spaceAfter=12,
            alignment=1,  # Center
        )
        self.styles.add(title_style)

        header_style = ParagraphStyle(
            "SectionHeader",
            parent=self.styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#1E40AF"),
            spaceAfter=10,
            spaceBefore=10,
        )
        self.styles.add(header_style)

    def export_student_schedule(
        self,
        student_name: str,
        group_name: str,
        entries: List[Dict[str, Any]],
        heatmap: Dict,
    ) -> bytes:
        """Export student weekly schedule to PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            topMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
            leftMargin=0.5 * inch,
        )

        story = []

        # Title
        title = Paragraph(f"Student Schedule - {student_name}", self.styles["CustomTitle"])
        story.append(title)

        # Info box
        info_data = [
            ["Student Name", student_name],
            ["Group", group_name],
            ["Classes", str(len(entries))],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
        info_table = Table(info_data, colWidths=[2 * inch, 4 * inch])
        info_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E0E7FF")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ])
        )
        story.append(info_table)
        story.append(Spacer(1, 0.3 * inch))

        # Schedule table
        story.append(Paragraph("Weekly Schedule", self.styles["SectionHeader"]))

        if entries:
            schedule_data = [
                ["Day", "Period", "Course", "Lecturer", "Room", "Time", "Duration"],
            ]

            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
            for entry in entries:
                day_name = days[entry.get("day", 0)] if entry.get("day", 0) < 5 else "N/A"
                schedule_data.append([
                    day_name,
                    str(entry.get("period", "")),
                    entry.get("course_code", "N/A"),
                    entry.get("lecturer_name", "N/A")[:15],
                    entry.get("room_name", "N/A"),
                    f"{entry.get('start_time', '')}-{entry.get('end_time', '')}",
                    f"{entry.get('duration_hours', '')}h",
                ])

            schedule_table = Table(schedule_data, colWidths=[1.2*inch, 0.8*inch, 1.2*inch, 1.2*inch, 1.0*inch, 1.2*inch, 0.8*inch])
            schedule_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                ])
            )
            story.append(schedule_table)
        else:
            story.append(Paragraph("No classes scheduled.", self.styles["Normal"]))

        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def export_lecturer_dashboard(
        self,
        lecturer_name: str,
        total_hours: float,
        course_count: int,
        session_count: int,
        sessions: List[Dict[str, Any]],
    ) -> bytes:
        """Export lecturer dashboard to PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            topMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
            leftMargin=0.5 * inch,
        )

        story = []

        # Title
        title = Paragraph(f"Lecturer Dashboard - {lecturer_name}", self.styles["CustomTitle"])
        story.append(title)

        # Stats box
        stats_data = [
            ["Total Hours", str(total_hours)],
            ["Courses", str(course_count)],
            ["Sessions", str(session_count)],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
        stats_table = Table(stats_data, colWidths=[2 * inch, 4 * inch])
        stats_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DBEAFE")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ])
        )
        story.append(stats_table)
        story.append(Spacer(1, 0.3 * inch))

        # Sessions table
        story.append(Paragraph("Teaching Schedule", self.styles["SectionHeader"]))

        if sessions:
            sessions_data = [
                ["Course", "Group", "Students", "Room", "Timeslot", "Duration"],
            ]

            for session in sessions:
                sessions_data.append([
                    session.get("course_code", "N/A"),
                    session.get("group_name", "N/A")[:10],
                    str(session.get("student_count", "")),
                    session.get("room_name", "N/A"),
                    f"TS{session.get('timeslot_id', '')}",
                    f"{session.get('duration_hours', '')}h",
                ])

            sessions_table = Table(sessions_data, colWidths=[1.5*inch, 1.2*inch, 0.9*inch, 1.2*inch, 1.2*inch, 0.9*inch])
            sessions_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0EA5E9")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F9FF")]),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                ])
            )
            story.append(sessions_table)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def export_master_pulse(
        self,
        workspace_name: str,
        total_rooms: int,
        avg_utilization: float,
        rooms_data: List[Dict[str, Any]],
    ) -> bytes:
        """Export admin master pulse view to PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            topMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
            leftMargin=0.5 * inch,
        )

        story = []

        # Title
        title = Paragraph(f"Master Pulse - {workspace_name}", self.styles["CustomTitle"])
        story.append(title)

        # Overview
        overview_data = [
            ["Total Rooms", str(total_rooms)],
            ["Avg Utilization", f"{avg_utilization:.1f}%"],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
        overview_table = Table(overview_data, colWidths=[2 * inch, 4 * inch])
        overview_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#FEE2E2")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ])
        )
        story.append(overview_table)
        story.append(Spacer(1, 0.3 * inch))

        # Rooms grid
        story.append(Paragraph("Room Utilization", self.styles["SectionHeader"]))

        if rooms_data:
            rooms_table_data = [
                ["Room", "Building", "Capacity", "Utilization", "Status"],
            ]

            for room in rooms_data:
                util = room.get("utilization_percent", 0)
                status = "Underutilized" if util < 30 else "Good" if util < 80 else "Over-booked"

                rooms_table_data.append([
                    room.get("room_name", "N/A")[:12],
                    room.get("building", "N/A")[:10],
                    str(room.get("capacity", "")),
                    f"{util:.1f}%",
                    status,
                ])

            rooms_table = Table(rooms_table_data, colWidths=[1.2*inch, 1.2*inch, 1.0*inch, 1.2*inch, 1.4*inch])
            rooms_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F97316")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF7ED")]),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                ])
            )
            story.append(rooms_table)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()


# ═════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═════════════════════════════════════════════════════════════

class ExcelExporter:
    """
    Generate Excel timetable reports.
    Supports: student schedule, lecturer dashboard, admin master pulse.
    """

    def __init__(self):
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def export_student_schedule(
        self,
        student_name: str,
        group_name: str,
        entries: List[Dict[str, Any]],
        heatmap: Dict,
    ) -> bytes:
        """Export student schedule to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Schedule"

        # Title and metadata
        ws["A1"] = "Student Schedule"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:G1")

        ws["A2"] = "Student Name:"
        ws["B2"] = student_name
        ws["A3"] = "Group:"
        ws["B3"] = group_name
        ws["A4"] = "Classes:"
        ws["B4"] = len(entries)
        ws["A5"] = "Generated:"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Headers
        headers = ["Day", "Period", "Course", "Lecturer", "Room", "Time", "Duration"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border

        # Data rows
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        for row, entry in enumerate(entries, 8):
            day_name = days[entry.get("day", 0)] if entry.get("day", 0) < 5 else "N/A"

            ws.cell(row=row, column=1).value = day_name
            ws.cell(row=row, column=2).value = entry.get("period", "")
            ws.cell(row=row, column=3).value = entry.get("course_code", "N/A")
            ws.cell(row=row, column=4).value = entry.get("lecturer_name", "N/A")
            ws.cell(row=row, column=5).value = entry.get("room_name", "N/A")
            ws.cell(row=row, column=6).value = f"{entry.get('start_time', '')}-{entry.get('end_time', '')}"
            ws.cell(row=row, column=7).value = f"{entry.get('duration_hours', '')}h"

            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 10

        # Heatmap sheet
        heatmap_ws = wb.create_sheet("Heatmap")
        heatmap_ws["A1"] = "Utilization Heatmap"
        heatmap_ws["A1"].font = Font(size=12, bold=True)

        row = 3
        for ts_id, util in heatmap.items():
            heatmap_ws.cell(row=row, column=1).value = f"Timeslot {ts_id}"
            heatmap_ws.cell(row=row, column=2).value = util
            heatmap_ws.cell(row=row, column=2).number_format = "0.0%"
            row += 1

        # Export to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def export_lecturer_dashboard(
        self,
        lecturer_name: str,
        total_hours: float,
        course_count: int,
        session_count: int,
        sessions: List[Dict[str, Any]],
    ) -> bytes:
        """Export lecturer dashboard to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard"

        # Title and stats
        ws["A1"] = f"Lecturer Dashboard - {lecturer_name}"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        ws.merge_cells("A1:F1")

        ws["A2"] = "Total Hours:"
        ws["B2"] = total_hours
        ws["A3"] = "Courses:"
        ws["B3"] = course_count
        ws["A4"] = "Sessions:"
        ws["B4"] = session_count
        ws["A5"] = "Generated:"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Headers
        headers = ["Course", "Group", "Students", "Room", "Timeslot", "Duration"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
            cell.border = self.thin_border

        # Data rows
        for row, session in enumerate(sessions, 8):
            ws.cell(row=row, column=1).value = session.get("course_code", "N/A")
            ws.cell(row=row, column=2).value = session.get("group_name", "N/A")
            ws.cell(row=row, column=3).value = session.get("student_count", "")
            ws.cell(row=row, column=4).value = session.get("room_name", "N/A")
            ws.cell(row=row, column=5).value = f"TS{session.get('timeslot_id', '')}"
            ws.cell(row=row, column=6).value = f"{session.get('duration_hours', '')}h"

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 14

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def export_master_pulse(
        self,
        workspace_name: str,
        total_rooms: int,
        avg_utilization: float,
        rooms_data: List[Dict[str, Any]],
    ) -> bytes:
        """Export master pulse to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Pulse"

        # Title
        ws["A1"] = f"Master Pulse - {workspace_name}"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        ws.merge_cells("A1:E1")

        ws["A2"] = "Total Rooms:"
        ws["B2"] = total_rooms
        ws["A3"] = "Avg Utilization:"
        ws["B3"] = avg_utilization
        ws["B3"].number_format = "0.0%"

        # Headers
        headers = ["Room", "Building", "Capacity", "Utilization", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
            cell.border = self.thin_border

        # Data rows
        for row, room in enumerate(rooms_data, 6):
            util = room.get("utilization_percent", 0)
            status = "Underutilized" if util < 30 else "Good" if util < 80 else "Over-booked"

            ws.cell(row=row, column=1).value = room.get("room_name", "N/A")
            ws.cell(row=row, column=2).value = room.get("building", "N/A")
            ws.cell(row=row, column=3).value = room.get("capacity", "")
            ws.cell(row=row, column=4).value = util / 100.0
            ws.cell(row=row, column=4).number_format = "0.0%"
            ws.cell(row=row, column=5).value = status

            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal="center" if col != 1 else "left")

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
